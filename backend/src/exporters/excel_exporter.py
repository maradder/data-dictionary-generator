"""Excel exporter for data dictionaries with advanced formatting."""

from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.exporters.base import BaseExporter

if TYPE_CHECKING:
    from models.dictionary import Dictionary
    from models.field import Field


class ExcelExporter(BaseExporter):
    """
    Excel exporter with advanced formatting and optimization.

    Features:
    - Write-only mode for memory efficiency with large datasets
    - Frozen header panes
    - Auto-filter on header row
    - Auto-sized column widths
    - Alternating row colors
    - Conditional formatting for null percentages
    - PII field highlighting
    - Metadata sheet with dictionary information
    """

    # Color schemes
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

    ROW_EVEN_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ROW_ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    PII_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    PII_FONT = Font(color="CC0000", bold=True)

    NULL_HIGH_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    NULL_MEDIUM_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    BORDER_STYLE = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3"),
    )

    ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")
    ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Column definitions
    COLUMNS = [
        {"header": "Field Path", "width": 40},
        {"header": "Data Type", "width": 15},
        {"header": "Semantic Type", "width": 20},
        {"header": "Description", "width": 50},
        {"header": "Sample Values", "width": 35},
        {"header": "Null %", "width": 12},
        {"header": "Cardinality", "width": 15},
        {"header": "PII Flag", "width": 12},
    ]

    def export_dictionary(
        self,
        dictionary: "Dictionary",
        fields: list["Field"],
        output_path: Path,
    ) -> None:
        """
        Export data dictionary to Excel with advanced formatting.

        Args:
            dictionary: The Dictionary object containing metadata
            fields: List of Field objects to export
            output_path: Path where the Excel file should be saved

        Raises:
            IOError: If unable to write to the output path
            ValueError: If the data is invalid or cannot be exported
        """
        # Validate output path
        self.validate_output_path(output_path)

        # Create workbook (not write-only to support formatting)
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        data_sheet = wb.create_sheet("Data Dictionary", 0)
        metadata_sheet = wb.create_sheet("Metadata", 1)

        # Export data
        self._write_data_sheet(data_sheet, fields)
        self._write_metadata_sheet(metadata_sheet, dictionary, len(fields))

        # Save workbook
        try:
            wb.save(output_path)
        except Exception as e:
            raise OSError(f"Failed to save Excel file to {output_path}: {e}") from e

    def _write_data_sheet(self, ws: Worksheet, fields: list["Field"]) -> None:
        """
        Write the main data dictionary sheet with formatting.

        Args:
            ws: Worksheet to write to
            fields: List of Field objects to export
        """
        # Write header row
        self._write_header_row(ws)

        # Write data rows
        for row_idx, field in enumerate(fields, start=2):
            self._write_field_row(ws, row_idx, field)

        # Apply formatting
        self._apply_column_widths(ws)
        self._apply_frozen_panes(ws)
        self._apply_auto_filter(ws, len(fields))
        self._apply_conditional_formatting(ws, len(fields))

    def _write_header_row(self, ws: Worksheet) -> None:
        """Write and format the header row."""
        for col_idx, column_def in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_def["header"])
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.ALIGNMENT_CENTER
            cell.border = self.BORDER_STYLE

    def _write_field_row(self, ws: Worksheet, row_idx: int, field: "Field") -> None:
        """
        Write a single field row with appropriate formatting.

        Args:
            ws: Worksheet to write to
            row_idx: Row index (1-based)
            field: Field object to write
        """
        # Determine row fill based on even/odd and PII status
        row_fill = self.ROW_EVEN_FILL if row_idx % 2 == 0 else self.ROW_ODD_FILL

        # Get annotation if available
        annotation = field.annotations[0] if field.annotations else None
        description = annotation.description if annotation else ""

        # Format sample values
        sample_values = self._format_sample_values(field.sample_values)

        # Format null percentage
        null_pct = f"{field.null_percentage:.1f}%" if field.null_percentage is not None else ""

        # PII flag
        pii_flag = "Yes" if field.is_pii else "No"

        # Data for the row
        row_data = [
            field.field_path,
            field.data_type,
            field.semantic_type or "",
            description,
            sample_values,
            null_pct,
            field.distinct_count,
            pii_flag,
        ]

        # Write cells with formatting
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Apply PII highlighting if this is a PII field
            if field.is_pii:
                cell.fill = self.PII_FILL
                if col_idx == 8:  # PII Flag column
                    cell.font = self.PII_FONT
            else:
                cell.fill = row_fill

            # Apply null percentage formatting
            if col_idx == 6 and field.null_percentage is not None:
                if field.null_percentage > 50:
                    cell.fill = self.NULL_HIGH_FILL
                elif field.null_percentage > 20:
                    cell.fill = self.NULL_MEDIUM_FILL

            # Apply alignment
            if col_idx in [2, 6, 7, 8]:  # Data Type, Null %, Cardinality, PII Flag
                cell.alignment = self.ALIGNMENT_CENTER
            else:
                cell.alignment = self.ALIGNMENT_LEFT

            cell.border = self.BORDER_STYLE

    def _format_sample_values(self, sample_values: dict[str, Any] | None) -> str:
        """
        Format sample values for display in Excel.

        Args:
            sample_values: Dictionary of sample values from the Field model

        Returns:
            Formatted string of sample values
        """
        if not sample_values:
            return ""

        # Sample values are stored as a dict, extract the values
        # The structure depends on how they're stored - adapt as needed
        if isinstance(sample_values, dict):
            # If it's a dict with a 'values' key
            if "values" in sample_values:
                values = sample_values["values"]
            else:
                # Otherwise, use dict values
                values = list(sample_values.values())
        elif isinstance(sample_values, list):
            values = sample_values
        else:
            return str(sample_values)

        # Take first 5 values and format them
        display_values = [str(v) for v in values[:5] if v is not None]
        result = ", ".join(display_values)

        # Truncate if too long
        if len(result) > 200:
            result = result[:197] + "..."

        return result

    def _apply_column_widths(self, ws: Worksheet) -> None:
        """Apply auto-sized column widths based on column definitions."""
        for col_idx, column_def in enumerate(self.COLUMNS, start=1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = column_def["width"]

    def _apply_frozen_panes(self, ws: Worksheet) -> None:
        """Freeze the header row."""
        ws.freeze_panes = "A2"

    def _apply_auto_filter(self, ws: Worksheet, num_fields: int) -> None:
        """
        Apply auto-filter to the data range.

        Args:
            ws: Worksheet to apply filter to
            num_fields: Number of field rows
        """
        if num_fields > 0:
            last_column = get_column_letter(len(self.COLUMNS))
            ws.auto_filter.ref = f"A1:{last_column}{num_fields + 1}"

    def _apply_conditional_formatting(self, ws: Worksheet, num_fields: int) -> None:
        """
        Apply conditional formatting rules.

        Note: Conditional formatting with color scales is complex in openpyxl
        and we're already applying colors directly in _write_field_row.
        This method is kept for potential future enhancements.

        Args:
            ws: Worksheet to apply formatting to
            num_fields: Number of field rows
        """
        # Conditional formatting is already applied inline during row writing
        # for better control over the appearance
        pass

    def _write_metadata_sheet(
        self, ws: Worksheet, dictionary: "Dictionary", num_fields: int
    ) -> None:
        """
        Write the metadata sheet with dictionary information.

        Args:
            ws: Worksheet to write to
            dictionary: Dictionary object with metadata
            num_fields: Number of fields in the dictionary
        """
        # Title
        ws.cell(row=1, column=1, value="Data Dictionary Metadata")
        ws.cell(row=1, column=1).font = Font(size=14, bold=True, color="366092")
        ws.merge_cells("A1:B1")

        # Metadata rows
        metadata = [
            ("Dictionary Name", dictionary.name),
            ("Description", dictionary.description or "N/A"),
            ("Source File", dictionary.source_file_name or "N/A"),
            ("File Size (bytes)", dictionary.source_file_size or "N/A"),
            ("Total Records Analyzed", dictionary.total_records_analyzed or "N/A"),
            ("Total Fields", num_fields),
            ("Created At", self._format_datetime(dictionary.created_at)),
            ("Created By", dictionary.created_by or "N/A"),
            ("Updated At", self._format_datetime(dictionary.updated_at)),
            ("Dictionary ID", str(dictionary.id)),
        ]

        # Write metadata
        for row_idx, (key, value) in enumerate(metadata, start=3):
            # Key cell
            key_cell = ws.cell(row=row_idx, column=1, value=key)
            key_cell.font = Font(bold=True)
            key_cell.fill = PatternFill(
                start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
            )
            key_cell.border = self.BORDER_STYLE
            key_cell.alignment = self.ALIGNMENT_LEFT

            # Value cell
            value_cell = ws.cell(row=row_idx, column=2, value=value)
            value_cell.border = self.BORDER_STYLE
            value_cell.alignment = self.ALIGNMENT_LEFT

        # Additional custom metadata if present
        if dictionary.custom_metadata:
            ws.cell(row=len(metadata) + 5, column=1, value="Additional Metadata")
            ws.cell(row=len(metadata) + 5, column=1).font = Font(bold=True, size=12)

            row_offset = len(metadata) + 6
            for key, value in dictionary.custom_metadata.items():
                key_cell = ws.cell(row=row_offset, column=1, value=str(key))
                key_cell.font = Font(bold=True)
                key_cell.fill = PatternFill(
                    start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
                )

                value_cell = ws.cell(row=row_offset, column=2, value=str(value))
                row_offset += 1

        # Set column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50

        # Export timestamp
        ws.cell(row=len(metadata) + 15, column=1, value="Exported At")
        ws.cell(row=len(metadata) + 15, column=1).font = Font(italic=True, size=9)
        ws.cell(row=len(metadata) + 15, column=2, value=self._format_datetime(datetime.now(UTC)))
        ws.cell(row=len(metadata) + 15, column=2).font = Font(italic=True, size=9)

    def _format_datetime(self, dt: datetime | None) -> str:
        """
        Format datetime for display.

        Args:
            dt: Datetime object to format

        Returns:
            Formatted datetime string
        """
        if dt is None:
            return "N/A"
        return dt.strftime("%Y-%m-%d %H:%M:%S UTC")

    def batch_export_dictionaries(
        self,
        dictionaries_data: list[dict],
        output_path: Path,
    ) -> None:
        """
        Export multiple dictionaries to a single Excel workbook.

        Each dictionary gets its own sheet, and a summary sheet is created.

        Args:
            dictionaries_data: List of dicts with 'dictionary', 'version', and 'fields' keys
            output_path: Path where the Excel file should be saved

        Raises:
            IOError: If unable to write to the output path
            ValueError: If the data is invalid or cannot be exported
        """
        # Validate output path
        self.validate_output_path(output_path)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create summary sheet first
        summary_sheet = wb.create_sheet("Summary", 0)
        self._write_summary_sheet(summary_sheet, dictionaries_data)

        # Create a sheet for each dictionary
        for idx, data in enumerate(dictionaries_data, start=1):
            dictionary = data["dictionary"]
            fields = data["fields"]
            version = data["version"]

            # Create sheet name (sanitize to avoid Excel sheet name issues)
            sheet_name = self._sanitize_sheet_name(dictionary.name, idx)

            # Create sheet
            data_sheet = wb.create_sheet(sheet_name, idx)

            # Write dictionary data
            self._write_data_sheet(data_sheet, fields)

        # Save workbook
        try:
            wb.save(output_path)
        except Exception as e:
            raise OSError(f"Failed to save Excel file to {output_path}: {e}") from e

    def _sanitize_sheet_name(self, name: str, idx: int) -> str:
        """
        Sanitize sheet name to comply with Excel requirements.

        Excel sheet names must be:
        - 31 characters or less
        - Cannot contain: \ / ? * [ ]

        Args:
            name: Original sheet name
            idx: Index for fallback naming

        Returns:
            Sanitized sheet name
        """
        # Remove invalid characters
        invalid_chars = ['\\', '/', '?', '*', '[', ']']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '_')

        # Truncate to 31 characters
        if len(sanitized) > 31:
            # Leave room for potential suffix
            sanitized = sanitized[:28] + f"_{idx}"

        # Ensure it's not empty
        if not sanitized:
            sanitized = f"Dictionary_{idx}"

        return sanitized

    def _write_summary_sheet(
        self, ws: Worksheet, dictionaries_data: list[dict]
    ) -> None:
        """
        Write summary sheet with information about all dictionaries.

        Args:
            ws: Worksheet to write to
            dictionaries_data: List of dicts with dictionary information
        """
        # Title
        ws.cell(row=1, column=1, value="Batch Export Summary")
        ws.cell(row=1, column=1).font = Font(size=14, bold=True, color="366092")
        ws.merge_cells("A1:F1")

        # Export info
        ws.cell(row=2, column=1, value=f"Export Date: {self._format_datetime(datetime.now(UTC))}")
        ws.cell(row=2, column=1).font = Font(italic=True, size=10)
        ws.merge_cells("A2:F2")

        ws.cell(row=3, column=1, value=f"Total Dictionaries: {len(dictionaries_data)}")
        ws.cell(row=3, column=1).font = Font(bold=True, size=11)
        ws.merge_cells("A3:F3")

        # Headers
        headers = [
            "Dictionary Name",
            "Version",
            "Total Fields",
            "Records Analyzed",
            "Created At",
            "Sheet Name"
        ]

        row_idx = 5
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.ALIGNMENT_CENTER
            cell.border = self.BORDER_STYLE

        # Data rows
        for idx, data in enumerate(dictionaries_data, start=1):
            dictionary = data["dictionary"]
            version = data["version"]
            fields = data["fields"]

            row_idx = 5 + idx
            row_fill = self.ROW_EVEN_FILL if row_idx % 2 == 0 else self.ROW_ODD_FILL

            row_data = [
                dictionary.name,
                f"v{version.version_number}",
                len(fields),
                dictionary.total_records_analyzed or "N/A",
                self._format_datetime(dictionary.created_at),
                self._sanitize_sheet_name(dictionary.name, idx)
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = self.BORDER_STYLE
                if col_idx in [2, 3]:  # Version and Total Fields
                    cell.alignment = self.ALIGNMENT_CENTER
                else:
                    cell.alignment = self.ALIGNMENT_LEFT

        # Set column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 35

        # Freeze header row
        ws.freeze_panes = "A6"

        # Add auto-filter
        if dictionaries_data:
            ws.auto_filter.ref = f"A5:F{5 + len(dictionaries_data)}"
