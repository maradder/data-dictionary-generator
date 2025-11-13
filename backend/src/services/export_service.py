"""
Export Service - Business logic for exporting data dictionaries.

This module provides the ExportService class for exporting dictionaries
to various formats (Excel, CSV, JSON, etc.).
"""

import logging
import tempfile
from pathlib import Path
from typing import Any
from uuid import UUID

from sqlalchemy.orm import Session

from src.core.exceptions import (
    ExportError,
    NotFoundError,
    ValidationError,
)
from src.exporters.excel_exporter import ExcelExporter
from src.models.dictionary import Dictionary
from src.models.field import Field
from src.models.version import Version

# Get logger
logger = logging.getLogger(__name__)
audit_logger = logging.getLogger("audit")


class ExportService:
    """
    Service for exporting data dictionaries.

    Handles exporting dictionaries to various formats including Excel,
    with support for version-specific exports.
    """

    def __init__(self, db: Session):
        """
        Initialize ExportService.

        Args:
            db: SQLAlchemy database session
        """
        self.db = db
        self.excel_exporter = ExcelExporter()

        logger.info("ExportService initialized")

    def get_dictionary_name(self, dictionary_id: UUID) -> str:
        """
        Get dictionary name by ID.

        Args:
            dictionary_id: Dictionary UUID

        Returns:
            str: Dictionary name

        Raises:
            NotFoundError: If dictionary not found
        """
        dictionary = (
            self.db.query(Dictionary)
            .filter(Dictionary.id == str(dictionary_id))
            .first()
        )

        if not dictionary:
            raise NotFoundError(f"Dictionary not found: {dictionary_id}")

        return dictionary.name

    def get_version_number(
        self, dictionary_id: UUID, version_id: UUID | None = None
    ) -> int:
        """
        Get version number by ID or latest if not specified.

        Args:
            dictionary_id: Dictionary UUID
            version_id: Optional specific version UUID

        Returns:
            int: Version number

        Raises:
            NotFoundError: If version not found
        """
        if version_id:
            version = (
                self.db.query(Version)
                .filter(Version.id == str(version_id), Version.dictionary_id == str(dictionary_id))
                .first()
            )
        else:
            # Get latest version
            version = (
                self.db.query(Version)
                .filter(Version.dictionary_id == str(dictionary_id))
                .order_by(Version.version_number.desc())
                .first()
            )

        if not version:
            raise NotFoundError("Version not found")

        return version.version_number

    def export_to_json(
        self,
        dictionary_id: UUID,
        version_id: UUID | None = None,
        include_statistics: bool = True,
        include_annotations: bool = True,
        include_pii_info: bool = True,
    ) -> dict[str, Any]:
        """
        Export dictionary to JSON format.

        Args:
            dictionary_id: Dictionary UUID
            version_id: Optional specific version ID (defaults to latest)
            include_statistics: Include statistical data
            include_annotations: Include annotations
            include_pii_info: Include PII information

        Returns:
            dict: Complete dictionary data as JSON

        Raises:
            NotFoundError: If dictionary or version not found
        """
        # Get dictionary
        dictionary = (
            self.db.query(Dictionary)
            .filter(Dictionary.id == str(dictionary_id))
            .first()
        )

        if not dictionary:
            raise NotFoundError(f"Dictionary not found: {dictionary_id}")

        # Get version
        if version_id:
            version = (
                self.db.query(Version)
                .filter(Version.id == str(version_id), Version.dictionary_id == str(dictionary_id))
                .first()
            )
        else:
            version = (
                self.db.query(Version)
                .filter(Version.dictionary_id == str(dictionary_id))
                .order_by(Version.version_number.desc())
                .first()
            )

        if not version:
            raise NotFoundError(f"No versions found for dictionary {dictionary_id}")

        # Get fields
        fields = (
            self.db.query(Field)
            .filter(Field.version_id == version.id)
            .order_by(Field.position)
            .all()
        )

        # Build JSON response
        result = {
            "dictionary": {
                "id": str(dictionary.id),
                "name": dictionary.name,
                "description": dictionary.description,
                "created_at": dictionary.created_at.isoformat(),
                "created_by": dictionary.created_by,
            },
            "version": {
                "id": str(version.id),
                "version_number": version.version_number,
                "created_at": version.created_at.isoformat(),
                "created_by": version.created_by,
            },
            "fields": [],
        }

        # Add fields
        for field in fields:
            field_data = {
                "field_path": field.field_path,
                "field_name": field.field_name,
                "data_type": field.data_type,
                "semantic_type": field.semantic_type,
                "is_nullable": field.is_nullable,
                "is_array": field.is_array,
            }

            if include_statistics:
                field_data["statistics"] = {
                    "null_count": field.null_count,
                    "null_percentage": float(field.null_percentage) if field.null_percentage else None,
                    "total_count": field.total_count,
                    "distinct_count": field.distinct_count,
                    "min_value": float(field.min_value) if field.min_value else None,
                    "max_value": float(field.max_value) if field.max_value else None,
                    "mean_value": float(field.mean_value) if field.mean_value else None,
                }

            if include_pii_info:
                field_data["pii"] = {
                    "is_pii": field.is_pii,
                    "pii_type": field.pii_type,
                }

            if include_annotations and field.annotation:
                field_data["annotation"] = {
                    "description": field.annotation.description,
                    "business_name": field.annotation.business_name,
                    "is_ai_generated": field.annotation.is_ai_generated,
                }

            result["fields"].append(field_data)

        return result

    def export_to_excel(
        self,
        dictionary_id: UUID,
        version_id: UUID | None = None,
        include_statistics: bool = True,
        include_annotations: bool = True,
        include_pii_info: bool = True,
        output_path: Path | None = None,
        exported_by: str | None = None,
    ) -> Path:
        """
        Export dictionary to Excel format.

        Args:
            dictionary_id: Dictionary UUID
            version_id: Optional specific version ID (defaults to latest)
            include_statistics: Include statistical data
            include_annotations: Include annotations
            include_pii_info: Include PII information
            output_path: Optional path where Excel file should be saved (creates temp file if None)
            exported_by: User performing the export

        Returns:
            Path to created Excel file

        Raises:
            NotFoundError: If dictionary or version not found
            ValidationError: If output path is invalid
            ExportError: If export fails
        """
        logger.info(
            f"Exporting dictionary {dictionary_id} to Excel",
            extra={
                "dictionary_id": str(dictionary_id),
                "output_path": str(output_path) if output_path else "temp",
                "version_id": str(version_id) if version_id else "latest",
            },
        )

        # Create temp file if no output path provided
        if output_path is None:
            temp_file = tempfile.NamedTemporaryFile(
                delete=False, suffix=".xlsx", mode="wb"
            )
            output_path = Path(temp_file.name)
            temp_file.close()
        else:
            # Validate output path
            if output_path.suffix.lower() not in [".xlsx", ".xls"]:
                raise ValidationError("Output file must have .xlsx or .xls extension")

            # Ensure parent directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)

        # Get dictionary
        dictionary = (
            self.db.query(Dictionary)
            .filter(Dictionary.id == str(dictionary_id))
            .first()
        )

        if not dictionary:
            raise NotFoundError(
                f"Dictionary not found: {dictionary_id}",
                details={"dictionary_id": str(dictionary_id)},
            )

        # Get version
        if version_id:
            version = (
                self.db.query(Version)
                .filter(Version.id == str(version_id), Version.dictionary_id == str(dictionary_id))
                .first()
            )
            if not version:
                raise NotFoundError(
                    f"Version not found: {version_id}",
                    details={"version_id": str(version_id)},
                )
        else:
            # Get latest version
            version = (
                self.db.query(Version)
                .filter(Version.dictionary_id == str(dictionary_id))
                .order_by(Version.version_number.desc())
                .first()
            )
            if not version:
                raise NotFoundError(
                    f"No versions found for dictionary {dictionary_id}",
                    details={"dictionary_id": str(dictionary_id)},
                )

        # Get fields for version
        fields = (
            self.db.query(Field)
            .filter(Field.version_id == str(version.id))
            .order_by(Field.position)
            .all()
        )

        if not fields:
            logger.warning(
                f"No fields found for version {version.id}",
                extra={"version_id": str(version.id)},
            )

        # Perform export
        try:
            self.excel_exporter.export_dictionary(
                dictionary=dictionary,
                fields=fields,
                output_path=output_path,
            )

            logger.info(
                f"Excel export completed successfully: {output_path}",
                extra={
                    "dictionary_id": str(dictionary_id),
                    "version_id": str(version.id),
                    "num_fields": len(fields),
                },
            )

        except Exception as e:
            logger.error(f"Excel export failed: {e}", exc_info=True)
            raise ExportError(
                f"Failed to export dictionary to Excel: {str(e)}",
                details={
                    "dictionary_id": str(dictionary_id),
                    "output_path": str(output_path),
                },
            )

        # Audit log
        audit_logger.info(
            "Dictionary exported to Excel",
            extra={
                "action": "export_to_excel",
                "dictionary_id": str(dictionary_id),
                "version_id": str(version.id),
                "version_number": version.version_number,
                "output_path": str(output_path),
                "exported_by": exported_by,
                "num_fields": len(fields),
            },
        )

        return output_path

    def export_version_comparison(
        self,
        dictionary_id: UUID,
        version_1_number: int,
        version_2_number: int,
        output_path: Path | None = None,
        exported_by: str | None = None,
    ) -> Path:
        """
        Export version comparison to Excel.

        Creates an Excel file showing differences between two versions.

        Args:
            dictionary_id: Dictionary UUID
            version_1_number: First version number
            version_2_number: Second version number
            output_path: Optional path where Excel file should be saved (creates temp file if None)
            exported_by: User performing the export

        Returns:
            Path to created Excel file

        Raises:
            NotFoundError: If dictionary or versions not found
            ValidationError: If output path is invalid
            ExportError: If export fails
        """
        from services.version_service import VersionService

        # Create temp file if no output path provided
        if output_path is None:
            temp_file = tempfile.NamedTemporaryFile(
                delete=False, suffix=".xlsx", mode="wb"
            )
            output_path = Path(temp_file.name)
            temp_file.close()

        # Get comparison data from version service
        version_service = VersionService(db=self.db)
        comparison_data = version_service.compare_versions(
            dictionary_id=dictionary_id,
            version_1_number=version_1_number,
            version_2_number=version_2_number,
        )
        logger.info(
            f"Exporting version comparison for dictionary {dictionary_id}",
            extra={
                "dictionary_id": str(dictionary_id),
                "v1": version_1_number,
                "v2": version_2_number,
            },
        )

        # Validate output path
        if output_path.suffix.lower() not in [".xlsx", ".xls"]:
            raise ValidationError("Output file must have .xlsx or .xls extension")

        # Ensure parent directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            from openpyxl import Workbook

            wb = Workbook()
            wb.remove(wb.active)

            # Create summary sheet
            summary_sheet = wb.create_sheet("Comparison Summary", 0)
            self._write_comparison_summary(summary_sheet, comparison_data)

            # Create changes sheet
            changes_sheet = wb.create_sheet("Changes Detail", 1)
            self._write_changes_detail(changes_sheet, comparison_data)

            # Save workbook
            wb.save(output_path)

            logger.info(
                f"Version comparison export completed: {output_path}",
                extra={
                    "dictionary_id": str(dictionary_id),
                    "v1": version_1_number,
                    "v2": version_2_number,
                },
            )

        except Exception as e:
            logger.error(f"Version comparison export failed: {e}", exc_info=True)
            raise ExportError(
                f"Failed to export version comparison: {str(e)}",
                details={
                    "dictionary_id": str(dictionary_id),
                    "v1": version_1_number,
                    "v2": version_2_number,
                },
            ) from e

        # Audit log
        audit_logger.info(
            "Version comparison exported",
            extra={
                "action": "export_version_comparison",
                "dictionary_id": str(dictionary_id),
                "v1": version_1_number,
                "v2": version_2_number,
                "output_path": str(output_path),
                "exported_by": exported_by,
            },
        )

        return output_path

    def _write_comparison_summary(self, ws, comparison_data: dict) -> None:
        """
        Write comparison summary sheet.

        Args:
            ws: Worksheet to write to
            comparison_data: Comparison results
        """
        from openpyxl.styles import Font

        # Title
        ws.cell(row=1, column=1, value="Version Comparison Summary")
        ws.cell(row=1, column=1).font = Font(size=14, bold=True, color="366092")

        # Version info
        v1_info = comparison_data.get("version_1", {})
        v2_info = comparison_data.get("version_2", {})

        ws.cell(row=3, column=1, value="Version 1")
        ws.cell(row=3, column=1).font = Font(bold=True)
        ws.cell(row=3, column=2, value=f"v{v1_info.get('version_number', 'N/A')}")
        ws.cell(row=4, column=1, value="Created")
        ws.cell(row=4, column=2, value=v1_info.get("created_at", "N/A"))

        ws.cell(row=6, column=1, value="Version 2")
        ws.cell(row=6, column=1).font = Font(bold=True)
        ws.cell(row=6, column=2, value=f"v{v2_info.get('version_number', 'N/A')}")
        ws.cell(row=7, column=1, value="Created")
        ws.cell(row=7, column=2, value=v2_info.get("created_at", "N/A"))

        # Summary statistics
        summary = comparison_data.get("summary", {})

        ws.cell(row=9, column=1, value="Change Summary")
        ws.cell(row=9, column=1).font = Font(bold=True)

        stats = [
            ("Fields Added", summary.get("fields_added", 0)),
            ("Fields Removed", summary.get("fields_removed", 0)),
            ("Fields Modified", summary.get("fields_modified", 0)),
            ("Breaking Changes", summary.get("breaking_changes", 0)),
            ("Total Fields (v1)", summary.get("total_fields_v1", 0)),
            ("Total Fields (v2)", summary.get("total_fields_v2", 0)),
        ]

        for i, (label, value) in enumerate(stats, start=10):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=1).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)

            # Highlight breaking changes
            if label == "Breaking Changes" and value > 0:
                ws.cell(row=i, column=2).font = Font(color="CC0000", bold=True)

        # Set column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

    def _write_changes_detail(self, ws, comparison_data: dict) -> None:
        """
        Write changes detail sheet.

        Args:
            ws: Worksheet to write to
            comparison_data: Comparison results
        """
        from openpyxl.styles import Alignment, Font, PatternFill

        # Header colors
        HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        HEADER_FONT = Font(color="FFFFFF", bold=True)
        ADDED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        REMOVED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        MODIFIED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        BREAKING_FONT = Font(color="CC0000", bold=True)

        # Headers
        headers = [
            "Change Type",
            "Field Path",
            "Breaking",
            "Old Type",
            "New Type",
            "Changes",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write changes
        changes = comparison_data.get("changes", [])
        row_idx = 2

        for change in changes:
            change_type = change.get("change_type", "")
            field_path = change.get("field_path", "")
            is_breaking = change.get("is_breaking", False)

            v1_data = change.get("version_1_data", {})
            v2_data = change.get("version_2_data", {})

            old_type = v1_data.get("data_type", "N/A")
            new_type = v2_data.get("data_type", "N/A")
            changes_list = change.get("changes", [])

            # Write row
            ws.cell(row=row_idx, column=1, value=change_type.upper())
            ws.cell(row=row_idx, column=2, value=field_path)
            ws.cell(row=row_idx, column=3, value="Yes" if is_breaking else "No")
            ws.cell(row=row_idx, column=4, value=old_type)
            ws.cell(row=row_idx, column=5, value=new_type)
            ws.cell(row=row_idx, column=6, value="; ".join(changes_list))

            # Color based on change type
            fill = None
            if change_type == "added":
                fill = ADDED_FILL
            elif change_type == "removed":
                fill = REMOVED_FILL
            elif change_type == "modified":
                fill = MODIFIED_FILL

            if fill:
                for col_idx in range(1, 7):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

            # Highlight breaking changes
            if is_breaking:
                ws.cell(row=row_idx, column=3).font = BREAKING_FONT

            row_idx += 1

        # Set column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 50

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add auto-filter
        if changes:
            ws.auto_filter.ref = f"A1:F{row_idx - 1}"

    def batch_export_to_excel(
        self,
        dictionary_ids: list[UUID],
        include_statistics: bool = True,
        include_annotations: bool = True,
        include_pii_info: bool = True,
        output_path: Path | None = None,
        exported_by: str | None = None,
    ) -> Path:
        """
        Export multiple dictionaries to a single Excel workbook.

        Each dictionary gets its own sheet with the latest version.
        Creates a summary sheet with metadata for all dictionaries.

        Args:
            dictionary_ids: List of dictionary UUIDs to export
            include_statistics: Include statistical data
            include_annotations: Include annotations
            include_pii_info: Include PII information
            output_path: Optional path where Excel file should be saved (creates temp file if None)
            exported_by: User performing the export

        Returns:
            Path to created Excel file

        Raises:
            NotFoundError: If any dictionary not found
            ValidationError: If no dictionaries provided or output path is invalid
            ExportError: If export fails
        """
        logger.info(
            f"Batch exporting {len(dictionary_ids)} dictionaries to Excel",
            extra={
                "num_dictionaries": len(dictionary_ids),
                "output_path": str(output_path) if output_path else "temp",
            },
        )

        # Validate input
        if not dictionary_ids:
            raise ValidationError("At least one dictionary ID is required")

        # Create temp file if no output path provided
        if output_path is None:
            temp_file = tempfile.NamedTemporaryFile(
                delete=False, suffix=".xlsx", mode="wb"
            )
            output_path = Path(temp_file.name)
            temp_file.close()
        else:
            # Validate output path
            if output_path.suffix.lower() not in [".xlsx", ".xls"]:
                raise ValidationError("Output file must have .xlsx or .xls extension")

            # Ensure parent directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)

        # Collect data for all dictionaries
        dictionaries_data = []
        for dictionary_id in dictionary_ids:
            # Get dictionary
            dictionary = (
                self.db.query(Dictionary)
                .filter(Dictionary.id == str(dictionary_id))
                .first()
            )

            if not dictionary:
                raise NotFoundError(
                    f"Dictionary not found: {dictionary_id}",
                    details={"dictionary_id": str(dictionary_id)},
                )

            # Get latest version
            version = (
                self.db.query(Version)
                .filter(Version.dictionary_id == str(dictionary_id))
                .order_by(Version.version_number.desc())
                .first()
            )

            if not version:
                raise NotFoundError(
                    f"No versions found for dictionary {dictionary_id}",
                    details={"dictionary_id": str(dictionary_id)},
                )

            # Get fields for version
            fields = (
                self.db.query(Field)
                .filter(Field.version_id == str(version.id))
                .order_by(Field.position)
                .all()
            )

            dictionaries_data.append({
                "dictionary": dictionary,
                "version": version,
                "fields": fields,
            })

        # Perform batch export
        try:
            self.excel_exporter.batch_export_dictionaries(
                dictionaries_data=dictionaries_data,
                output_path=output_path,
            )

            logger.info(
                f"Batch Excel export completed successfully: {output_path}",
                extra={
                    "num_dictionaries": len(dictionary_ids),
                    "output_path": str(output_path),
                },
            )

        except Exception as e:
            logger.error(f"Batch Excel export failed: {e}", exc_info=True)
            raise ExportError(
                f"Failed to batch export dictionaries to Excel: {str(e)}",
                details={
                    "num_dictionaries": len(dictionary_ids),
                    "output_path": str(output_path),
                },
            )

        # Audit log
        audit_logger.info(
            "Dictionaries batch exported to Excel",
            extra={
                "action": "batch_export_to_excel",
                "dictionary_ids": [str(d_id) for d_id in dictionary_ids],
                "num_dictionaries": len(dictionary_ids),
                "output_path": str(output_path),
                "exported_by": exported_by,
            },
        )

        return output_path
