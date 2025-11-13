"""Import service for rehydrating database from XLSX exports."""
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

import openpyxl
from sqlalchemy.orm import Session

from src.core.config import settings
from src.core.exceptions import ValidationError
from src.models.annotation import Annotation
from src.models.dictionary import Dictionary
from src.models.field import Field
from src.models.version import Version

logger = logging.getLogger(__name__)


class ImportService:
    """Service for importing dictionaries from XLSX files."""

    def __init__(self, db: Session):
        """
        Initialize import service.

        Args:
            db: SQLAlchemy database session
        """
        self.db = db

    def import_from_excel(
        self,
        file_path: Path | str,
        conflict_mode: str = "skip",
        imported_by: str | None = None,
    ) -> dict[str, Any]:
        """
        Import dictionaries from an XLSX export file.

        Args:
            file_path: Path to the XLSX file
            conflict_mode: How to handle conflicts ("skip", "overwrite", "fail")
            imported_by: User performing the import

        Returns:
            dict: Import results including counts and any errors

        Raises:
            ValidationError: If file format is invalid
            FileNotFoundError: If file doesn't exist
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"Import file not found: {file_path}")

        if file_path.suffix.lower() not in [".xlsx", ".xls"]:
            raise ValidationError("Import file must be an Excel file (.xlsx or .xls)")

        logger.info(f"Starting import from {file_path}")

        results = {
            "status": "success",
            "file": str(file_path),
            "imported_by": imported_by,
            "started_at": datetime.now(timezone.utc),
            "dictionaries_imported": 0,
            "versions_imported": 0,
            "fields_imported": 0,
            "annotations_imported": 0,
            "errors": [],
            "warnings": [],
            "skipped": [],
        }

        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)

            # Determine import type (single dictionary or batch)
            if "Summary" in wb.sheetnames:
                # Batch export
                results.update(self._import_batch_export(wb, conflict_mode, imported_by))
            elif "Data Dictionary" in wb.sheetnames and "Metadata" in wb.sheetnames:
                # Single dictionary export
                results.update(
                    self._import_single_dictionary(wb, conflict_mode, imported_by)
                )
            else:
                raise ValidationError(
                    "Invalid export file format. Expected 'Data Dictionary' and 'Metadata' "
                    "sheets or a 'Summary' sheet for batch exports."
                )

            results["completed_at"] = datetime.now(timezone.utc)
            logger.info(
                f"Import completed: {results['dictionaries_imported']} dictionaries, "
                f"{results['fields_imported']} fields"
            )

        except Exception as e:
            logger.error(f"Import failed: {e}", exc_info=True)
            results["status"] = "failed"
            results["error"] = str(e)
            results["completed_at"] = datetime.now(timezone.utc)

        return results

    def _import_single_dictionary(
        self, wb: openpyxl.Workbook, conflict_mode: str, imported_by: str | None
    ) -> dict[str, Any]:
        """Import a single dictionary from workbook."""
        results = {
            "dictionaries_imported": 0,
            "versions_imported": 0,
            "fields_imported": 0,
            "annotations_imported": 0,
            "errors": [],
            "warnings": [],
            "skipped": [],
        }

        metadata_sheet = wb["Metadata"]
        data_sheet = wb["Data Dictionary"]

        # Parse metadata
        metadata = self._parse_metadata_sheet(metadata_sheet)

        # Validate dictionary ID
        dict_id = metadata.get("dictionary_id")
        if not dict_id:
            # Generate new ID if not present
            dict_id = str(uuid4())
            results["warnings"].append("No dictionary_id found, generated new ID")

        # Check for conflicts
        existing_dict = self.db.query(Dictionary).filter(Dictionary.id == str(dict_id)).first()

        if existing_dict:
            if conflict_mode == "skip":
                results["skipped"].append(
                    {"dictionary_id": dict_id, "reason": "Already exists"}
                )
                return results
            elif conflict_mode == "fail":
                raise ValidationError(
                    f"Dictionary {dict_id} already exists and conflict_mode is 'fail'"
                )
            elif conflict_mode == "overwrite":
                # Delete existing dictionary (cascade will handle relations)
                self.db.delete(existing_dict)
                self.db.flush()
                results["warnings"].append(
                    f"Overwrote existing dictionary {dict_id}"
                )

        # Create dictionary
        dictionary = Dictionary(
            id=dict_id,
            name=metadata.get("name", "Imported Dictionary"),
            description=metadata.get("description"),
            source_file_name=metadata.get("source_file"),
            source_file_size=metadata.get("source_file_size"),
            total_records_analyzed=metadata.get("total_records"),
            created_at=metadata.get("created_at", datetime.now(timezone.utc)),
            created_by=imported_by or metadata.get("created_by", "import"),
            updated_at=datetime.now(timezone.utc),
            custom_metadata=metadata.get("metadata"),
        )

        self.db.add(dictionary)
        self.db.flush()
        results["dictionaries_imported"] += 1

        # Create version
        version_id = metadata.get("version_id", str(uuid4()))
        version = Version(
            id=version_id,
            dictionary_id=dict_id,
            version_number=metadata.get("version_number", 1),
            schema_hash=metadata.get("schema_hash", "imported"),
            created_at=metadata.get("version_created_at", datetime.now(timezone.utc)),
            created_by=imported_by or metadata.get("created_by", "import"),
            notes=metadata.get("version_notes"),
            processing_stats=metadata.get("processing_stats"),
        )

        self.db.add(version)
        self.db.flush()
        results["versions_imported"] += 1

        # Parse and create fields
        fields_data = self._parse_data_sheet(data_sheet)

        for field_data in fields_data:
            field = self._create_field_from_data(field_data, version_id)
            self.db.add(field)
            results["fields_imported"] += 1

            # Create annotations if present
            if field_data.get("description"):
                annotation = Annotation(
                    id=str(uuid4()),
                    field_id=field.id,
                    description=field_data["description"],
                    business_name=field_data.get("business_name"),
                    is_ai_generated=field_data.get("is_ai_generated", False),
                    created_at=datetime.now(timezone.utc),
                    created_by=imported_by or "import",
                    updated_at=datetime.now(timezone.utc),
                )
                self.db.add(annotation)
                results["annotations_imported"] += 1

        # Commit in batches for better performance
        if settings.is_sqlite and results["fields_imported"] > settings.batch_commit_size:
            self.db.commit()
            logger.debug(f"Batch commit: {results['fields_imported']} fields")

        self.db.commit()
        return results

    def _import_batch_export(
        self, wb: openpyxl.Workbook, conflict_mode: str, imported_by: str | None
    ) -> dict[str, Any]:
        """Import multiple dictionaries from a batch export."""
        results = {
            "dictionaries_imported": 0,
            "versions_imported": 0,
            "fields_imported": 0,
            "annotations_imported": 0,
            "errors": [],
            "warnings": [],
            "skipped": [],
        }

        # Skip the Summary sheet, process each dictionary sheet
        for sheet_name in wb.sheetnames:
            if sheet_name == "Summary":
                continue

            try:
                # Each sheet is a dictionary
                # For batch exports, we'd need to parse each sheet differently
                # This is a simplified version - full implementation would parse
                # the actual batch export format
                results["warnings"].append(
                    f"Batch import for sheet '{sheet_name}' not fully implemented yet"
                )
            except Exception as e:
                results["errors"].append({"sheet": sheet_name, "error": str(e)})
                logger.error(f"Failed to import sheet {sheet_name}: {e}")

        return results

    def _parse_metadata_sheet(self, sheet) -> dict[str, Any]:
        """Parse the Metadata sheet to extract dictionary information."""
        metadata = {}

        # Metadata sheet has key-value pairs in columns A and B
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                key = str(row[0]).strip().lower().replace(" ", "_")
                value = row[1]

                # Handle special fields
                if key == "metadata" and isinstance(value, str):
                    try:
                        metadata["metadata"] = json.loads(value)
                    except json.JSONDecodeError:
                        metadata["metadata"] = {"raw": value}
                elif key in ["created_at", "version_created_at"] and isinstance(value, datetime):
                    metadata[key] = value
                else:
                    metadata[key] = value

        return metadata

    def _parse_data_sheet(self, sheet) -> list[dict[str, Any]]:
        """Parse the Data Dictionary sheet to extract field information."""
        fields = []

        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                # Normalize header names
                header = str(cell.value).strip().lower().replace(" ", "_")
                headers.append(header)

        # Parse data rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue

            field_data = {}
            for idx, value in enumerate(row):
                if idx < len(headers):
                    header = headers[idx]
                    # Handle special fields
                    if header == "sample_values" and isinstance(value, str):
                        try:
                            field_data[header] = json.loads(value)
                        except json.JSONDecodeError:
                            field_data[header] = {"values": [value] if value else []}
                    elif header in ["is_nullable", "is_array", "is_pii", "is_ai_generated"]:
                        field_data[header] = bool(value) if value else False
                    else:
                        field_data[header] = value

            fields.append(field_data)

        return fields

    def _create_field_from_data(self, field_data: dict[str, Any], version_id: str) -> Field:
        """Create a Field object from parsed data."""
        return Field(
            id=field_data.get("id", str(uuid4())),
            version_id=version_id,
            field_path=field_data.get("field_path", "unknown"),
            field_name=field_data.get("field_name", field_data.get("field_path", "unknown")),
            parent_path=field_data.get("parent_path"),
            nesting_level=int(field_data.get("nesting_level", 0)),
            data_type=field_data.get("data_type", "unknown"),
            semantic_type=field_data.get("semantic_type"),
            is_nullable=field_data.get("is_nullable", True),
            is_array=field_data.get("is_array", False),
            array_item_type=field_data.get("array_item_type"),
            sample_values=field_data.get("sample_values"),
            null_count=int(field_data.get("null_count", 0)) if field_data.get("null_count") else 0,
            null_percentage=float(field_data.get("null_percentage", 0)) if field_data.get("null_percentage") else None,
            total_count=int(field_data.get("total_count", 0)) if field_data.get("total_count") else 0,
            distinct_count=int(field_data.get("distinct_count", 0)) if field_data.get("distinct_count") else 0,
            cardinality_ratio=float(field_data.get("cardinality_ratio", 0)) if field_data.get("cardinality_ratio") else None,
            min_value=float(field_data.get("min_value")) if field_data.get("min_value") else None,
            max_value=float(field_data.get("max_value")) if field_data.get("max_value") else None,
            mean_value=float(field_data.get("mean_value")) if field_data.get("mean_value") else None,
            median_value=float(field_data.get("median_value")) if field_data.get("median_value") else None,
            std_dev=float(field_data.get("std_dev")) if field_data.get("std_dev") else None,
            percentile_25=float(field_data.get("percentile_25")) if field_data.get("percentile_25") else None,
            percentile_50=float(field_data.get("percentile_50")) if field_data.get("percentile_50") else None,
            percentile_75=float(field_data.get("percentile_75")) if field_data.get("percentile_75") else None,
            is_pii=field_data.get("is_pii", False),
            pii_type=field_data.get("pii_type"),
            confidence_score=float(field_data.get("confidence_score")) if field_data.get("confidence_score") else None,
            created_at=datetime.now(timezone.utc),
            position=int(field_data.get("position")) if field_data.get("position") else None,
        )

    def validate_import_file(self, file_path: Path | str) -> dict[str, Any]:
        """
        Validate an import file without importing it.

        Args:
            file_path: Path to the XLSX file

        Returns:
            dict: Validation results including structure check and potential issues
        """
        file_path = Path(file_path)
        validation = {
            "valid": True,
            "file": str(file_path),
            "format": None,
            "issues": [],
            "warnings": [],
            "summary": {},
        }

        try:
            if not file_path.exists():
                validation["valid"] = False
                validation["issues"].append("File not found")
                return validation

            wb = openpyxl.load_workbook(file_path, data_only=True)

            # Determine format
            if "Summary" in wb.sheetnames:
                validation["format"] = "batch_export"
            elif "Data Dictionary" in wb.sheetnames and "Metadata" in wb.sheetnames:
                validation["format"] = "single_dictionary"
                # Count fields
                data_sheet = wb["Data Dictionary"]
                field_count = sum(1 for row in data_sheet.iter_rows(min_row=2) if row[0].value)
                validation["summary"]["estimated_fields"] = field_count
            else:
                validation["valid"] = False
                validation["issues"].append("Unrecognized file format")

        except Exception as e:
            validation["valid"] = False
            validation["issues"].append(f"Failed to read file: {str(e)}")

        return validation
